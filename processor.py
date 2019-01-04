#!/env/lib/python3
# -*- coding: utf-8 -*-
import os
import csv
import requests
import shutil
import openpyxl
import asyncio
from number import getPhone
from bs4 import BeautifulSoup
from PyQt5.QtWidgets import QMainWindow

class Processor(object):
	
	def __init__(self):
		self.file_name = ''
		self.file = ''
		self.array_len = 0
		self.status = True
	
	def transliterate(self, name):
	
		slovar = {'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'e',
				'ж':'zh','з':'z','и':'i','й':'i','к':'k','л':'l','м':'m','н':'n',
				'о':'o','п':'p','р':'r','с':'s','т':'t','у':'u','ф':'f','х':'h',
				'ц':'c','ч':'cz','ш':'sh','щ':'scz','ъ':'','ы':'y','ь':'','э':'e',
				'ю':'u','я':'ya', 'А':'a','Б':'b','В':'v','Г':'g','Д':'d','Е':'e','Ё':'e',
				'Ж':'zh','З':'z','И':'i','Й':'i','К':'k','Л':'l','М':'m','Н':'n',
				'О':'o','П':'p','Р':'r','С':'s','Т':'t','У':'u','Ф':'F','х':'h',
				'Ц':'c','Ч':'cz','Ш':'sh','Щ':'scz','Ъ':'','Ы':'y','Ь':'','Э':'e',
				'Ю':'u','Я':'ya',',':'','?':'',' ':'_','~':'','!':'','@':'','#':'',
				'$':'','%':'','^':'','&':'','*':'','(':'',')':'','-':'','=':'','+':'',
				':':'',';':'','<':'','>':'','\'':'','"':'','\\':'','/':'','№':'',
				'[':'',']':'','{':'','}':'','ґ':'','ї':'', 'є':'','Ґ':'g','Ї':'i',
				'Є':'e'
			}
	
		for key in slovar:
			name = name.replace(key, slovar[key])
		return name
	
	def get_html(self, url):
		r = requests.get(url)
		return r.text
	
	def file_save_xlsx(self, data, i):
		workbook = openpyxl.load_workbook(self.file)
		sheet = workbook.active
		
		i+=1
		c1 = sheet.cell(row=i, column=1)
		c1.value = data['title']
		
		c2 = sheet.cell(row=i, column=2)
		c2.value = data['price']
		
		c3 = sheet.cell(row=i, column=3)
		c3.value = data['address']
		
		c4 = sheet.cell(row=i, column=4)
		c4.value = data['phone']

		c5 = sheet.cell(row=i, column=5)
		c5.value = data['url']

		workbook.save(self.file)
		

	async def get_page_data(self, html):
		soup = BeautifulSoup(html, 'lxml')
		try:
			ads = soup.find('div',
				class_='catalog-list').find_all('div',
				class_='item_table')
			
			self.status = True
		
		except AttributeError:
			self.status = False
			return False
		
		self.array_len = len(ads)
		
		directory = os.path.expanduser("~/Desktop/avito_parser/")
		
		if not os.path.exists(directory):
			os.makedirs(directory)
	
		self.file = directory + self.file_name + '.xlsx'
		
		book = openpyxl.Workbook(self.file)
		book.save(self.file)

		i = 0 
		for ad in ads:
			
			try:
				title = ad.find('div', class_='description').find('h3').text.strip()
			except:
				title = ''

			try:
				url = 'https://www.avito.ru' + ad.find('div', class_='description').find('h3').find('a').get('href')
				m_url = 'https://m.avito.ru/' + ad.find('div', class_='description').find('h3').find('a').get('href')
			except:
				url = ''

			try:
				price = ad.find('div', class_='about').text.strip()
			except:
				price = ''

			try:
				phone = getPhone(m_url)
			
			except:
				phone = ''

			try:
				address = ad.find('p', class_='address').text.strip()
			except:
				address = ''

			data = {
				'title': title,
				'price': price,
				'address': address,
				'phone': phone,
				'url': url
			}
			self.file_save_xlsx(data, i)
			i+=1
			await asyncio.sleep(0.002)
		
	async def result_hub(self, city, category, loop):
		self.file_name = city + '_' + category
		
		base_url = 'https://www.avito.ru/'
		part = '?p='
		
		for i in range(1, 3):
			url_gen = base_url + city + '/' + category + '/' + part + str(i)
			html = self.get_html(url_gen)
			loop.create_task(self.get_page_data(html))
		await asyncio.sleep(0.002)
